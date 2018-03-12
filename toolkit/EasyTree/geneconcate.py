#!/usr/bin/python
# -*- coding: UTF-8 -*-
# Filename: geneconcate
# Desc:  gene modify
# author: leo.liu
# date: 2018.2.28

from openpyxl import workbook
from openpyxl import load_workbook
import sys
import os
import codecs
#from openpyxl.write.excel import ExcelWriter

# 第一个需求
# 1. Datasheet中保存了Unicode，及其对应的碱基对LSU/ITS/.../..
# 2. 每个碱基对类型（LSU/ITS/.../..）都有一个同名文件对应，该文件保存 碱基对与编码的对应关系，同一个碱基对的编码长度是相同的
#
# 3. 按Datasheet中Unicode的顺序，及其对应的碱基对LSU/ITS/.../..进行重新生成，生成规则为：
# 	1）新生成的文件名以碱基对类型拼接（LSU+ITS+....txt）
# 	2）新生成的文件内容为：
# 		Unicode
# 		碱基对对应的编码连续拼接
# 		Datasheet中Unicode对应的碱基对不存在的，以其定长的n符号代替
#


# 实现方案
# 1. 加载Datasheet到数组中
# 2. 加载所有碱基对，{type,{key:value, k:v}} eg.  {LSU{KY212762:nnnnnnnn}}
# 3. 每个碱基对类型，添加一个{None:nnnnnnn}，其中nnnn的长度与其他基因相同
# 4. 顺序按行扫描Datasheet，拼接Unicode, gen1 gen2.....
# 5. 写入文本文件，名字为gen1+gen2+....

# 碱基对起始列
GEN_START = 4
# 结束列，从后数
GEN_END   = 2
# 分隔符
SPLITTER  = '\n'
# None str
NONE_STR = 'n'

DATA_PATH = U'./data/'
RST_PATH = U'./data/result/'
DATASHEET = DATA_PATH + u'Datasheet.xlsx'


def LoadDatasheet(filename):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    row = len(list(ws.rows))
    col = len(list(ws.columns)) # 最后两列不要了

    data = [[0 for i in range(col)] for i in range(row)]

    for r in range(1, row+1):
        for c in range(1, col+1):
            data[r-1][c-1] = ws.cell(row=r, column=c).value
    return data, row, col

def LoadGen(gentypes):
    # 第一维存基因类型
    # 第二维存碱基对（KTxxxxx: nnnnnnnnnnnn)
    allGen = {}
    for t in gentypes:
        filename = DATA_PATH + t + '.txt'
        # print filename
        with open(filename, 'r') as f:
            n = 0
            k = ""
            v = ""
            genLen = 0
            genpair = {}

            while True:
                n=n+1
                line = f.readline().strip()
                # print 'line ------', n, line
                if not line:
                    break

                if n % 2 == 1:
                    # 去掉行头的 > 符号
                    k = line[1:]
                    genpair[k] = ''
                else:
                    v = line
                    genpair[k] = v
                    genLen = len(v)
            #print genpair
            # nn = ""
            # print genLen
            # for i in range(0, genLen):
            #     nn = nn + 'n'
            genpair["None"] = NONE_STR*genLen #str("".join('n') for i in range(0, genLen))
            # print 'all gen:', t, '=', genpair
            allGen[t] = genpair
    # print '------------a--------a----a-'
    # print allGen
    return allGen

# dna/rna文件的组织方式不是按行，而是以>为开头标记，因此重写此方法
def LoadGen2(gentypes):
    # 第一维存基因类型
    # 第二维存碱基对（KTxxxxx: nnnnnnnnnnnn)
    #print 'Loading gen pairs...'
    allGen = {}
    for t in gentypes:
        filename = DATA_PATH + t + '.txt'
        # print filename
        with open(filename, 'r') as f:
            k = ''
            v = ''
            genLen = 0
            genpair = {}
            while True:
                line = f.readline().strip()
                # print line
                if not line:
                    break

                if line.startswith(">"):
                    k = line[1:]
                    v = ''
                    genpair[k] = ''
                else:
                    v += line + SPLITTER
                    genpair[k] = v
                    genLen = len(v)

            genpair["None"] = 'n'*genLen
            allGen[t] = genpair

        #print allGen
    return allGen


def genconcate(datasheet, gens):
    head = datasheet[0]
    data = datasheet[1:]
    rst = {}
    for row in data:
        uid = row[0]
        constr = ""
        for c in range(GEN_START, len(row)-GEN_END):
            #print (head[c], row[c])
            if row[c] is None:
                row[c] = 'None'
            constr += gens[head[c]][row[c]]
        rst[uid] = constr

    return rst

def writetofile(gencon, filename):
    with open(filename, 'wb') as f:
        keys = gencon.keys()
        keys.sort()
        for k in keys:
            v = gencon[k]
        #for (k, v) in gencon.items():
            f.write(k+'\n')
            f.write(v+'\n')

def do_main():
    # Load from Datasheet.xlsx
    print 'Loading dataset...'
    (datasheet, row, col) = LoadDatasheet(DATASHEET)
    gentypes = datasheet[0][4:col-GEN_END]
    # print gentypes
    outfile = RST_PATH
    for t in gentypes:
        outfile = outfile + t + '+'
    outfile = outfile[:len(outfile)-1] + '.txt'
    # 基因拼接
    print 'Concating...'
    writetofile(genconcate(datasheet, LoadGen2(gentypes)), outfile)
    print 'File saved in: ', outfile

    print 'Done.'
    return True

def main():
    if not os.path.exists(RST_PATH):
        os.mkdir(RST_PATH)
    do_main()

if __name__ =="__main__":
    main()
